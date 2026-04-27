from __future__ import annotations

from csv import reader as csv_reader
from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO, StringIO
from pathlib import Path
from typing import Sequence

import matplotlib
from openpyxl import load_workbook

matplotlib.use("Agg")

from matplotlib import pyplot as plt

SUPPORTED_PLOT_FILE_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}
CHART_TYPE_LABELS: dict[str, str] = {
    "scatter": "散點圖",
    "line": "折線圖",
    "bar": "長條圖",
}
MAX_PLOT_SELECT_OPTIONS = 25
MAX_PLOT_Y_SERIES = 5
PLOT_PREVIEW_ROW_COUNT = 6


class PlotPreviewError(ValueError):
    """Raised when the uploaded workbook cannot be parsed safely."""


class PlotRenderError(ValueError):
    """Raised when the workbook selection cannot be rendered into a chart."""


@dataclass(frozen=True)
class PlotColumnPreview:
    name: str
    inferred_type: str
    numeric: bool
    sample_values: tuple[str, ...] = ()


@dataclass(frozen=True)
class PlotSheetPreview:
    name: str
    row_count: int
    columns: tuple[PlotColumnPreview, ...]
    preview_rows: tuple[tuple[str, ...], ...] = ()
    data_rows: tuple[tuple[str, ...], ...] = ()
    has_header: bool = True


@dataclass(frozen=True)
class PlotWorkbookPreview:
    filename: str
    file_kind: str
    sheets: tuple[PlotSheetPreview, ...]


@dataclass
class PlotSelectionState:
    sheet_index: int = 0
    x_index: int = 0
    y_indices: list[int] = field(default_factory=list)
    chart_type: str = "scatter"
    title: str = ""
    x_axis_label: str = ""
    y_axis_label: str = ""


def is_supported_plot_file(filename: str) -> bool:
    return Path(str(filename or "")).suffix.lower() in SUPPORTED_PLOT_FILE_EXTENSIONS


def chart_type_label(chart_type: str) -> str:
    return CHART_TYPE_LABELS.get(str(chart_type or "").strip().lower(), "散點圖")


def parse_workbook_preview(filename: str, blob: bytes) -> PlotWorkbookPreview:
    suffix = Path(str(filename or "")).suffix.lower()
    if suffix not in SUPPORTED_PLOT_FILE_EXTENSIONS:
        raise PlotPreviewError("目前先支援 `.xlsx`、`.xlsm`、`.csv`。")
    if not blob:
        raise PlotPreviewError("檔案是空的，XE3 目前沒辦法讀這份資料。")

    if suffix == ".csv":
        rows = _parse_csv_rows(blob)
        sheet = _sheet_preview_from_rows("CSV", rows)
        return PlotWorkbookPreview(filename=filename or "data.csv", file_kind="csv", sheets=(sheet,))

    workbook = load_workbook(BytesIO(blob), read_only=True, data_only=True)
    sheets: list[PlotSheetPreview] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        sheets.append(_sheet_preview_from_rows(sheet_name, rows))

    if not sheets:
        raise PlotPreviewError("這份 Excel 裡目前沒有可讀的工作表。")

    return PlotWorkbookPreview(filename=filename or "workbook.xlsx", file_kind="excel", sheets=tuple(sheets))


def default_plot_selection(preview: PlotWorkbookPreview) -> PlotSelectionState:
    state = PlotSelectionState()
    clamp_plot_selection(preview, state)
    return state


def clamp_plot_selection(preview: PlotWorkbookPreview, state: PlotSelectionState) -> PlotSelectionState:
    if not preview.sheets:
        raise PlotPreviewError("這份資料目前沒有可選欄位。")

    state.sheet_index = max(0, min(int(state.sheet_index or 0), len(preview.sheets) - 1))
    sheet = preview.sheets[state.sheet_index]
    if not sheet.columns:
        state.x_index = 0
        state.y_indices = []
        state.chart_type = _normalized_chart_type(state.chart_type)
        return state

    visible_indexes = list(range(min(len(sheet.columns), MAX_PLOT_SELECT_OPTIONS)))
    if not visible_indexes:
        visible_indexes = [0]

    state.x_index = state.x_index if state.x_index in visible_indexes else visible_indexes[0]

    normalized_y = [idx for idx in state.y_indices if idx in visible_indexes and idx != state.x_index]
    if not normalized_y:
        numeric_indexes = [idx for idx in visible_indexes if sheet.columns[idx].numeric and idx != state.x_index]
        if numeric_indexes:
            normalized_y = numeric_indexes[:1]
        elif len(visible_indexes) > 1:
            fallback = visible_indexes[1] if visible_indexes[0] == state.x_index else visible_indexes[0]
            normalized_y = [fallback] if fallback != state.x_index else []
    state.y_indices = normalized_y[:MAX_PLOT_Y_SERIES]
    state.chart_type = _normalized_chart_type(state.chart_type)
    return state


def selected_sheet(preview: PlotWorkbookPreview, state: PlotSelectionState) -> PlotSheetPreview:
    return preview.sheets[max(0, min(state.sheet_index, len(preview.sheets) - 1))]


def selection_summary_text(preview: PlotWorkbookPreview, state: PlotSelectionState) -> str:
    sheet = selected_sheet(preview, state)
    x_name = sheet.columns[state.x_index].name if sheet.columns and state.x_index < len(sheet.columns) else "尚未選擇"
    y_names = [sheet.columns[idx].name for idx in state.y_indices if idx < len(sheet.columns)]
    y_label = "、".join(y_names) if y_names else "尚未選擇"
    x_samples = _selected_series_sample_text(sheet, state.x_index)
    y_sample_lines = _selected_y_sample_lines(sheet, state.y_indices)
    lines = [
        f"📁 工作表：**{sheet.name}**",
        f"📝 圖表標題：**{resolved_plot_title(preview, state)}**",
        f"↔️ X 軸名稱：**{resolved_x_axis_label(preview, state)}**",
        f"   • 對應資料欄：**{x_name}**{x_samples}",
        f"📊 Y 軸名稱：**{resolved_y_axis_label(preview, state)}**",
        f"   • 對應資料欄：**{y_label}**",
        f"📈 圖表：**{chart_type_label(state.chart_type)}**",
    ]
    if y_sample_lines:
        lines.extend(y_sample_lines)
    if not sheet.has_header:
        lines.extend(
            [
                "",
                "ℹ️ 這份資料看起來沒有明確欄名，XE3 先用 `預設X軸 / 預設Y軸 / 預設Y軸2...` 來幫你預覽。",
            ]
        )
    alignment_lines = _selection_alignment_preview_lines(preview, sheet, state)
    if alignment_lines:
        lines.extend(["", "🔍 對齊預覽", *alignment_lines])
    lines.extend(
        [
            "",
            "這組設定已經可以直接拿去出圖了。你如果只想先確認方向，也可以按 `完成設定` 再檢查一次。",
        ]
    )
    return "\n".join(lines)


def build_plot_template_csv() -> bytes:
    rows = [
        "X,Y",
        "0,0.139",
        "20,0.150",
        "40,0.161",
        "60,0.163",
        "80,0.178",
        "100,0.187",
    ]
    return ("\n".join(rows) + "\n").encode("utf-8")


def resolved_plot_title(preview: PlotWorkbookPreview, state: PlotSelectionState) -> str:
    custom = str(state.title or "").strip()
    if custom:
        return custom
    sheet = selected_sheet(preview, state)
    return sheet.name or "未命名圖表"


def resolved_x_axis_label(preview: PlotWorkbookPreview, state: PlotSelectionState) -> str:
    custom = str(state.x_axis_label or "").strip()
    if custom:
        return custom
    sheet = selected_sheet(preview, state)
    if sheet.columns and state.x_index < len(sheet.columns):
        return sheet.columns[state.x_index].name
    return "X"


def resolved_y_axis_label(preview: PlotWorkbookPreview, state: PlotSelectionState) -> str:
    custom = str(state.y_axis_label or "").strip()
    if custom:
        return custom
    sheet = selected_sheet(preview, state)
    if state.y_indices:
        names = [sheet.columns[idx].name for idx in state.y_indices if idx < len(sheet.columns)]
        if names:
            return "、".join(names)
    return "Y"


def render_plot_png(preview: PlotWorkbookPreview, state: PlotSelectionState) -> bytes:
    _configure_matplotlib_fonts()
    sheet = selected_sheet(preview, state)
    if not sheet.columns or not sheet.data_rows:
        raise PlotRenderError("這個工作表目前沒有可畫圖的資料列。")
    if not state.y_indices:
        raise PlotRenderError("請先至少選一個 Y 軸資料欄，再讓 XE3 幫你出圖。")

    x_label = resolved_x_axis_label(preview, state)
    y_label = resolved_y_axis_label(preview, state)
    title = resolved_plot_title(preview, state)
    chart_type = _normalized_chart_type(state.chart_type)
    x_mode = _detect_x_mode(sheet, state.x_index)
    if chart_type == "scatter" and x_mode == "category":
        raise PlotRenderError("散點圖目前需要可轉成數值或日期的 X 軸資料。")

    prepared_rows = _prepared_plot_rows(sheet, state, x_mode=x_mode)
    if not prepared_rows:
        raise PlotRenderError("找不到可用的數值資料。請確認 X / Y 欄位是不是選對了。")

    series_indexes = [idx for idx in state.y_indices if idx < len(sheet.columns)]
    if not series_indexes:
        raise PlotRenderError("目前沒有可用的 Y 軸資料欄。")

    fig, ax = plt.subplots(figsize=(8.8, 5.4), dpi=160)
    try:
        if chart_type == "bar":
            _draw_bar_chart(ax, sheet, prepared_rows, series_indexes, x_mode=x_mode)
        else:
            _draw_xy_chart(ax, chart_type, sheet, prepared_rows, series_indexes)

        ax.set_title(title)
        ax.set_xlabel(x_label)
        ax.set_ylabel(y_label if len(series_indexes) == 1 else y_label or "數值")
        ax.grid(True, linestyle="--", linewidth=0.7, alpha=0.25)
        if chart_type == "bar":
            plt.setp(ax.get_xticklabels(), rotation=20, ha="right")
        elif x_mode == "category":
            plt.setp(ax.get_xticklabels(), rotation=20, ha="right")
        if len(series_indexes) > 1:
            ax.legend(frameon=False)
        fig.tight_layout()
        buffer = BytesIO()
        fig.savefig(buffer, format="png", bbox_inches="tight")
        return buffer.getvalue()
    finally:
        plt.close(fig)


def apply_axis_label_mapping_hint(preview: PlotWorkbookPreview, state: PlotSelectionState) -> tuple[PlotSelectionState, str | None]:
    sheet = selected_sheet(preview, state)
    if sheet.has_header or len(sheet.columns) != 2 or len(state.y_indices) != 1:
        return state, None

    x_role = _axis_label_role(state.x_axis_label)
    y_role = _axis_label_role(state.y_axis_label)
    if not x_role and not y_role:
        return state, None

    current_x = state.x_index
    current_y = state.y_indices[0]
    direct_score = _mapping_score(sheet, current_x, current_y)
    swapped_score = _mapping_score(sheet, current_y, current_x)

    wants_independent_x = x_role == "independent"
    wants_dependent_y = y_role == "dependent"
    wants_dependent_x = x_role == "dependent"
    wants_independent_y = y_role == "independent"

    should_prefer_swapped = False
    if wants_independent_x and wants_dependent_y and swapped_score > direct_score:
        should_prefer_swapped = True
    elif wants_dependent_x and wants_independent_y and direct_score > swapped_score:
        should_prefer_swapped = True

    if not should_prefer_swapped:
        return state, None

    new_state = PlotSelectionState(
        sheet_index=state.sheet_index,
        x_index=current_y,
        y_indices=[current_x],
        chart_type=state.chart_type,
        title=state.title,
        x_axis_label=state.x_axis_label,
        y_axis_label=state.y_axis_label,
    )
    clamp_plot_selection(preview, new_state)
    return new_state, "XE3 看你填的軸名像是在表示 `time / absorbance` 這種配對，所以幫你把 X/Y 的資料欄位對調成更合理的方向。"


def _parse_csv_rows(blob: bytes) -> list[list[object]]:
    encodings = ("utf-8-sig", "utf-8", "cp950", "big5")
    decoded = None
    for encoding in encodings:
        try:
            decoded = blob.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        decoded = blob.decode("latin-1")
    return [list(row) for row in csv_reader(StringIO(decoded))]


def _sheet_preview_from_rows(sheet_name: str, rows: Sequence[Sequence[object]]) -> PlotSheetPreview:
    normalized_rows = [list(row) for row in rows]
    if not normalized_rows:
        return PlotSheetPreview(name=sheet_name, row_count=0, columns=(), preview_rows=(), has_header=True)

    header_idx, headers, has_header = _choose_headers(normalized_rows)
    data_rows = [row for row in normalized_rows[header_idx + 1 :] if any(_clean_cell(value) for value in row)]
    columns = tuple(_build_column_preview(headers, data_rows, idx) for idx in range(len(headers)))
    all_rows = tuple(
        tuple(_clean_cell(_value_at(row, idx)) for idx in range(len(headers)))
        for row in data_rows
    )
    preview_rows = all_rows[:PLOT_PREVIEW_ROW_COUNT]
    return PlotSheetPreview(
        name=sheet_name,
        row_count=len(data_rows),
        columns=columns,
        preview_rows=preview_rows,
        data_rows=all_rows,
        has_header=has_header,
    )


def _choose_headers(rows: Sequence[Sequence[object]]) -> tuple[int, list[str], bool]:
    for idx, row in enumerate(rows[:10]):
        cleaned = [_clean_cell(value) for value in row]
        non_empty = [cell for cell in cleaned if cell]
        if len(non_empty) >= 2 and _looks_like_header_row(row):
            return idx, _dedupe_headers(cleaned), True

    width = max(len(row) for row in rows[:10]) if rows else 0
    return -1, _default_headers(width), False


def _dedupe_headers(headers: Sequence[str]) -> list[str]:
    cleaned: list[str] = []
    seen: dict[str, int] = {}
    width = max(len(headers), 1)
    for idx in range(width):
        raw = str(headers[idx] or "").strip() if idx < len(headers) else ""
        base = raw or f"欄位 {idx + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        cleaned.append(base if count == 0 else f"{base} ({count + 1})")
    return cleaned


def _default_headers(width: int) -> list[str]:
    labels: list[str] = []
    for idx in range(max(width, 1)):
        if idx == 0:
            labels.append("預設X軸")
        elif idx == 1:
            labels.append("預設Y軸")
        else:
            labels.append(f"預設Y軸{idx}")
    return labels


def _build_column_preview(headers: Sequence[str], data_rows: Sequence[Sequence[object]], index: int) -> PlotColumnPreview:
    values = [_value_at(row, index) for row in data_rows]
    cleaned = [_clean_cell(value) for value in values if _clean_cell(value)]
    sample_values = tuple(cleaned[:5])
    non_empty_values = [value for value in values if _clean_cell(value)]

    if not non_empty_values:
        return PlotColumnPreview(name=headers[index], inferred_type="空白欄位", numeric=False, sample_values=sample_values)

    date_like = sum(1 for value in non_empty_values if isinstance(value, (datetime, date, time)))
    numeric_like = sum(1 for value in non_empty_values if _is_numeric_cell(value))
    total = len(non_empty_values)

    if date_like == total:
        inferred = "日期 / 時間"
        numeric = False
    elif numeric_like == total:
        inferred = "數值"
        numeric = True
    elif numeric_like >= max(1, int(total * 0.7)):
        inferred = "多數是數值"
        numeric = True
    else:
        inferred = "文字 / 混合"
        numeric = False

    return PlotColumnPreview(name=headers[index], inferred_type=inferred, numeric=numeric, sample_values=sample_values)


def _value_at(row: Sequence[object], index: int) -> object:
    return row[index] if index < len(row) else None


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def _is_numeric_cell(value: object) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.strip())
            return True
        except ValueError:
            return False
    return False


def _looks_like_header_row(row: Sequence[object]) -> bool:
    non_empty = [value for value in row if _clean_cell(value)]
    if len(non_empty) < 2:
        return False
    numeric_like = sum(1 for value in non_empty if _is_numeric_cell(value))
    date_like = sum(1 for value in non_empty if isinstance(value, (datetime, date, time)))
    if numeric_like == len(non_empty) or date_like == len(non_empty):
        return False
    return numeric_like <= max(0, len(non_empty) // 2)


def _normalized_chart_type(chart_type: str) -> str:
    normalized = str(chart_type or "").strip().lower()
    return normalized if normalized in CHART_TYPE_LABELS else "scatter"


def _selected_series_sample_text(sheet: PlotSheetPreview, index: int) -> str:
    if index >= len(sheet.columns):
        return ""
    sample_values = [value for value in sheet.columns[index].sample_values[:4] if value]
    if not sample_values:
        return ""
    return f"（例：`{' / '.join(sample_values)}`）"


def _selected_y_sample_lines(sheet: PlotSheetPreview, y_indices: Sequence[int]) -> list[str]:
    lines: list[str] = []
    for position, index in enumerate(y_indices, start=1):
        if index >= len(sheet.columns):
            continue
        label = "Y" if position == 1 else f"Y{position}"
        name = sheet.columns[index].name
        suffix = _selected_series_sample_text(sheet, index)
        lines.append(f"   • {label}：**{name}**{suffix}")
    return lines


def _selection_alignment_preview_lines(preview: PlotWorkbookPreview, sheet: PlotSheetPreview, state: PlotSelectionState) -> list[str]:
    if not sheet.preview_rows:
        return []
    selected_indices = [state.x_index, *state.y_indices]
    selected_indices = [idx for idx in selected_indices if idx < len(sheet.columns)]
    if not selected_indices:
        return []

    lines: list[str] = []
    x_alias = _short_axis_alias(resolved_x_axis_label(preview, state), fallback="X")
    y_alias = _short_axis_alias(resolved_y_axis_label(preview, state), fallback="Y")
    for row_idx, row in enumerate(sheet.preview_rows[:4], start=1):
        parts: list[str] = []
        for position, idx in enumerate(selected_indices):
            value = row[idx] if idx < len(row) else ""
            if position == 0:
                label = x_alias
            else:
                label = y_alias if position == 1 else f"Y{position}"
            parts.append(f"{label}=`{value or '∅'}`")
        if parts:
            lines.append(f"• 第 {row_idx} 筆｜" + " ｜ ".join(parts))
    return lines


def _axis_label_role(label: str) -> str:
    lowered = str(label or "").strip().lower()
    if not lowered:
        return ""
    independent_tokens = ("time", "時間", "濃度", "conc", "concentration", "dose", "volume", "temp", "temperature")
    dependent_tokens = ("abs", "absorb", "absorbance", "od", "signal", "intensity", "response", "吸光", "吸收")
    if any(token in lowered for token in independent_tokens):
        return "independent"
    if any(token in lowered for token in dependent_tokens):
        return "dependent"
    return ""


def _mapping_score(sheet: PlotSheetPreview, x_index: int, y_index: int) -> int:
    return _independent_score(_column_preview_values(sheet, x_index)) + _dependent_score(_column_preview_values(sheet, y_index))


def _column_preview_values(sheet: PlotSheetPreview, index: int) -> list[str]:
    values: list[str] = []
    for row in sheet.preview_rows:
        if index < len(row):
            value = str(row[index] or "").strip()
            if value:
                values.append(value)
    return values


def _independent_score(values: Sequence[str]) -> int:
    floats = _coerce_float_list(values)
    if len(floats) < 2:
        return 0
    score = 0
    if all(b >= a for a, b in zip(floats, floats[1:])):
        score += 2
    integer_like = sum(1 for value in values if _looks_integer_like(value))
    if integer_like >= max(1, len(values) // 2):
        score += 1
    if len(set(values)) >= 4:
        score += 1
    return score


def _dependent_score(values: Sequence[str]) -> int:
    if not values:
        return 0
    score = 0
    decimal_like = sum(1 for value in values if "." in value)
    if decimal_like >= max(1, len(values) // 2):
        score += 2
    floats = _coerce_float_list(values)
    if len(floats) >= 2 and any(abs(b - a) < 5 for a, b in zip(floats, floats[1:])):
        score += 1
    return score


def _coerce_float_list(values: Sequence[str]) -> list[float]:
    floats: list[float] = []
    for value in values:
        try:
            floats.append(float(str(value).strip()))
        except ValueError:
            return []
    return floats


def _looks_integer_like(value: str) -> bool:
    try:
        parsed = float(str(value).strip())
    except ValueError:
        return False
    return parsed.is_integer()


def _configure_matplotlib_fonts() -> None:
    plt.rcParams["font.sans-serif"] = [
        "Noto Sans CJK TC",
        "Noto Sans CJK SC",
        "Noto Sans CJK JP",
        "Microsoft JhengHei",
        "PingFang TC",
        "WenQuanYi Zen Hei",
        "SimHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False


def _detect_x_mode(sheet: PlotSheetPreview, x_index: int) -> str:
    values = _column_full_values(sheet, x_index)
    if not values:
        return "category"
    if all(_try_float(value) is not None for value in values):
        return "numeric"
    if all(_try_datetime(value) is not None for value in values):
        return "datetime"
    return "category"


def _column_full_values(sheet: PlotSheetPreview, index: int) -> list[str]:
    values: list[str] = []
    for row in sheet.data_rows:
        if index < len(row):
            value = str(row[index] or "").strip()
            if value:
                values.append(value)
    return values


def _prepared_plot_rows(
    sheet: PlotSheetPreview,
    state: PlotSelectionState,
    *,
    x_mode: str,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    series_indexes = [idx for idx in state.y_indices if idx < len(sheet.columns)]
    for row in sheet.data_rows:
        x_text = str(row[state.x_index] if state.x_index < len(row) else "").strip()
        if not x_text:
            continue
        x_value = _coerce_x_value(x_text, x_mode)
        if x_value is None:
            continue

        y_values: list[float | None] = []
        has_any_y = False
        for y_index in series_indexes:
            y_text = str(row[y_index] if y_index < len(row) else "").strip()
            y_value = _try_float(y_text)
            y_values.append(y_value)
            has_any_y = has_any_y or y_value is not None
        if not has_any_y:
            continue

        rows.append(
            {
                "x_raw": x_text,
                "x_value": x_value,
                "y_values": y_values,
            }
        )
    return rows


def _coerce_x_value(value: str, mode: str) -> object | None:
    if mode == "numeric":
        return _try_float(value)
    if mode == "datetime":
        return _try_datetime(value)
    text = str(value).strip()
    return text or None


def _try_float(value: str | None) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _try_datetime(value: str | None) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    formats = (
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _draw_xy_chart(ax, chart_type: str, sheet: PlotSheetPreview, prepared_rows: Sequence[dict[str, object]], series_indexes: Sequence[int]) -> None:
    for position, y_index in enumerate(series_indexes):
        xs: list[object] = []
        ys: list[float] = []
        for row in prepared_rows:
            y_values = row["y_values"]
            y_value = y_values[position] if position < len(y_values) else None
            if y_value is None:
                continue
            xs.append(row["x_value"])
            ys.append(y_value)
        if not ys:
            continue

        label = sheet.columns[y_index].name
        if chart_type == "scatter":
            ax.scatter(xs, ys, s=38, alpha=0.9, label=label)
        else:
            ax.plot(xs, ys, marker="o", linewidth=2, markersize=4, alpha=0.95, label=label)


def _draw_bar_chart(ax, sheet: PlotSheetPreview, prepared_rows: Sequence[dict[str, object]], series_indexes: Sequence[int], *, x_mode: str) -> None:
    import numpy as np

    x_positions = np.arange(len(prepared_rows))
    labels = [str(row["x_raw"]) for row in prepared_rows]
    series_count = max(len(series_indexes), 1)
    width = 0.8 / series_count

    for position, y_index in enumerate(series_indexes):
        offsets = x_positions + (position - (series_count - 1) / 2) * width
        valid_positions: list[float] = []
        valid_values: list[float] = []
        for row_idx, row in enumerate(prepared_rows):
            y_values = row["y_values"]
            y_value = y_values[position] if position < len(y_values) else None
            if y_value is None:
                continue
            valid_positions.append(offsets[row_idx])
            valid_values.append(y_value)
        if valid_values:
            ax.bar(valid_positions, valid_values, width=width, alpha=0.9, label=sheet.columns[y_index].name)

    ax.set_xticks(x_positions)
    ax.set_xticklabels(labels)


def _short_axis_alias(label: str, *, fallback: str) -> str:
    cleaned = str(label or "").strip()
    if not cleaned:
        return fallback
    return cleaned[:12]
